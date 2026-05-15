from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from tools.pipeline.runner import run_pipeline

ROOT = Path(__file__).resolve().parents[1]
FIXTURE_CAT = ROOT / "tests" / "fixtures" / "minimal_catalog.json"


def _valid_eval() -> dict:
    return {
        "product_fit_score": 5,
        "product_fit_reasons": ["与数字表目录匹配"],
        "capability_score": 5,
        "capability_signals": ["证据显示有分销渠道"],
        "reputation_risk": {"facts": [], "concerns": [], "sources": []},
        "reputation_safety_score": 5,
        "buyer_seller_role": "buyer",
        "buyer_seller_reason": "证据显示其为欧盟地区分销商，采购万用表类产品。",
        "deal_recommendation": "high_intent",
        "next_action": "报价",
        "confidence": 0.8,
        "data_quality": "high",
        "citations": [{"claim": "匹配", "source_url": "", "source_snippet": "摘录"}],
        "catalog_version_note": "test-fixture",
    }


@patch("tools.pipeline.runner.run_llm_eval")
def test_mock_llm_writes_summary_detail_and_highlight(mock_llm, tmp_path: Path) -> None:
    mock_llm.return_value = _valid_eval()
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    df = pd.DataFrame(
        [
            {
                "company_name": "MockCo",
                "website": "",
                "evidence_paste": "We distribute digital multimeters in EU.",
            }
        ]
    )
    df.to_excel(inp, index=False, engine="openpyxl")

    run_pipeline(
        inp,
        out,
        dry_run=False,
        no_fetch=True,
        limit=1,
        catalog_path=FIXTURE_CAT,
        kb_path=ROOT / "product_kb" / "v1" / "kb.json",
    )

    summary = pd.read_excel(out, sheet_name="Summary", engine="openpyxl")
    assert "合作建议" in summary.columns
    assert summary.loc[0, "合作建议"] == "高意向跟进"
    assert summary.loc[0, "买方/卖方"] == "买方"
    assert "与数字表目录匹配" in str(summary.loc[0, "产品匹配说明"])
    detail = pd.read_excel(out, sheet_name="Detail", engine="openpyxl")
    assert len(detail) == 1 and "eval_json" in detail.columns

    wb = load_workbook(out)
    ws = wb["Summary"]
    first_cell = ws.cell(row=2, column=1)
    assert first_cell.fill.fill_type == "solid"


def test_missing_catalog_raises(tmp_path: Path) -> None:
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    df = pd.DataFrame([{"company_name": "X", "evidence_paste": "y"}])
    df.to_excel(inp, index=False, engine="openpyxl")
    with pytest.raises(FileNotFoundError):
        run_pipeline(
            inp,
            out,
            dry_run=False,
            no_fetch=True,
            limit=1,
            catalog_path=tmp_path / "does_not_exist.json",
        )


@patch("tools.pipeline.runner.run_llm_eval")
def test_stop_on_error_raises(mock_llm, tmp_path: Path) -> None:
    mock_llm.side_effect = RuntimeError("boom")
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    df = pd.DataFrame([{"company_name": "X", "evidence_paste": "some evidence text here"}])
    df.to_excel(inp, index=False, engine="openpyxl")
    with pytest.raises(RuntimeError, match="boom"):
        run_pipeline(
            inp,
            out,
            dry_run=False,
            no_fetch=True,
            limit=1,
            catalog_path=FIXTURE_CAT,
            stop_on_error=True,
        )
