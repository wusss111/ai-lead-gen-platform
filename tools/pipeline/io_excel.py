from __future__ import annotations

import difflib
import json
import logging
import re
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from tools.pipeline.paths import SCHEMA_EXCEL_IO

logger = logging.getLogger(__name__)

# Excel 单元格禁止的控制字符（XML 1.0 规范）
_EXCEL_ILLEGAL_RE = re.compile('[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def _sanitize_excel_text(val: Any) -> Any:
    """移除 Excel 单元格不接受的非法控制字符。"""
    if isinstance(val, str):
        return _EXCEL_ILLEGAL_RE.sub('', val)
    return val

# 中文表头 / 别名 → 规范 input 列名（大小写不敏感匹配走 canonical）
COLUMN_ALIASES_TO_CANON: dict[str, str] = {
    # 公司名称
    "客户名称": "company_name", "公司名称": "company_name", "客户公司": "company_name",
    "商户名称": "company_name", "客户": "company_name", "买方": "company_name",
    "卖方": "company_name", "公司": "company_name", "企业名称": "company_name",
    "公司名": "company_name", "厂商名称": "company_name", "供应商名称": "company_name",
    "company": "company_name", "company name": "company_name", "customer": "company_name",
    "customer name": "company_name", "client": "company_name", "client name": "company_name",
    "帐户名称": "company_name", "账号名称": "company_name", "buyer": "company_name",
    "supplier": "company_name", "business name": "company_name",
    # 网站
    "企业网站": "website", "网址": "website", "官网": "website", "网站": "website",
    "公司网址": "website", "公司网站": "website", "客户网址": "website",
    "企业网址": "website", "网站地址": "website", "网页": "website",
    "homepage": "website", "url": "website", "web": "website", "link": "website",
    "website": "website", "web site": "website", "domain": "website",
    # 国家
    "洲": "country_region", "国家地区": "country_region", "国家": "country_region",
    "地区": "country_region", "区域": "country_region", "所属洲": "country_region",
    "region": "country_region", "country": "country_region", "country/region": "country_region",
    "location": "country_region",
    # 备注
    "备注": "notes", "说明": "notes", "留言": "notes", "摘要": "notes",
    "remark": "notes", "remarks": "notes", "note": "notes",
    "description": "notes", "comment": "notes",
    # 证据粘贴
    "粘贴证据": "evidence_paste", "证据摘录": "evidence_paste",
    "补充说明": "evidence_paste", "额外信息": "evidence_paste",
    # 联系信息
    "联系地址": "contact_address", "联系人地址": "contact_address", "地址": "contact_address",
    "固定电话": "contact_phone", "电话": "contact_phone", "联系电话": "contact_phone",
    "手机": "contact_phone", "phone": "contact_phone", "tel": "contact_phone",
    "联系人邮箱": "contact_email", "邮箱": "contact_email", "邮件": "contact_email",
    "email": "contact_email", "e-mail": "contact_email", "电子邮箱": "contact_email",
    "联系人姓名": "contact_name", "联系人": "contact_name", "姓名": "contact_name",
    "contact person": "contact_name", "contact": "contact_name",
    # 目标产品
    "目标产品": "target_products", "产品": "target_products", "关注产品": "target_products",
    "意向产品": "target_products", "products": "target_products",
    # 优先级
    "优先级": "priority", "重要程度": "priority", "priority": "priority",
}

# 常见表头关键词 → 目标列（模糊匹配时优先使用）
_COLUMN_KEYWORDS: dict[str, str] = {
    "公司": "company_name", "客户": "company_name", "企业": "company_name",
    "网址": "website", "网站": "website", "官网": "website",
    "国家": "country_region", "洲": "country_region",
    "邮箱": "contact_email", "邮件": "contact_email",
    "电话": "contact_phone", "手机": "contact_phone", "phone": "contact_phone", "tel": "contact_phone", "mobile": "contact_phone",
    "联系人": "contact_name", "姓名": "contact_name",
    "地址": "contact_address",
    "备注": "notes", "说明": "notes", "留言": "notes",
    "产品": "target_products",
    "优先": "priority",
}

_NOTES_FROM_CONTACTS: list[tuple[str, str]] = [
    ("联系地址", "contact_address"),
    ("固定电话", "contact_phone"),
    ("联系人邮箱", "contact_email"),
    ("联系人姓名", "contact_name"),
]


# ============================================================
#  数据清洗与预处理
# ============================================================

def _is_header_row(row_values: list[str]) -> bool:
    """判断一行数据是否像表头（而非数据行）。"""
    text_count = 0
    url_count = 0
    numeric_count = 0
    for v in row_values:
        s = str(v).strip().lower()
        if not s or s in ("nan", "none", "null", "-", "n/a", "na"):
            continue
        if s.startswith(("http://", "https://")):
            url_count += 1
            continue
        # 纯数字/日期 → 不太可能是表头
        if re.match(r"^[\d.,/\-]+$", s):
            numeric_count += 1
            continue
        # 含中文/英文关键词 → 更像是列名
        if re.search(r"[一-鿿]", s) or re.search(r"[a-zA-Z]{2,}", s):
            text_count += 1
    # 如果大部分列是 URL 或数字，这行很可能是数据，不是表头
    total = len(row_values)
    if total == 0:
        return False
    header_ratio = text_count / max(total, 1)
    data_ratio = (url_count + numeric_count) / max(total, 1)
    return header_ratio > data_ratio


def _detect_header_row(df: pd.DataFrame) -> int:
    """检测哪一行是实际的表头，返回表头行号（0-based）。
    如果第一行不像表头但第二行也不像，认为没有表头。
    """
    max_check = min(5, len(df))
    for i in range(max_check):
        row_vals = [str(v) for v in df.iloc[i].values]
        if _is_header_row(row_vals):
            return i
    return -1  # 没有检测到表头


def _handle_merged_cells(df: pd.DataFrame) -> pd.DataFrame:
    """处理合并单元格：对数据行 NaN 进行 forward fill（跳过表头行避免串行）。"""
    out = df.copy()
    for col in out.columns:
        # 只处理数据行（跳过表头 row 0），避免表头值被 ffill 传播到数据
        if len(out) <= 1:
            continue
        data_rows = out[col].iloc[1:]
        if data_rows.isna().any():
            out[col].iloc[1:] = data_rows.ffill()
    return out


def _dedup_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove duplicate columns (keep first occurrence)."""
    cols = df.columns.tolist()
    seen: set[str] = set()
    keep: list[int] = []
    for i, c in enumerate(cols):
        if c not in seen:
            seen.add(c)
            keep.append(i)
    if len(keep) < len(cols):
        logger = logging.getLogger(__name__)
        logger.warning("Dropped %d duplicate columns, kept %d/%d", len(cols) - len(keep), len(keep), len(cols))
    return df.iloc[:, keep]


def _strip_empty_rows_cols(df: pd.DataFrame) -> pd.DataFrame:
    """删除全空的行和列。"""
    out = df.dropna(how="all").copy()
    out = out.dropna(axis=1, how="all")
    return out


def _standardize_contact_formats(df: pd.DataFrame) -> pd.DataFrame:
    """标准化手机号和邮箱格式。"""
    out = df.copy()

    # 邮箱标准化
    if "contact_email" in out.columns:
        def clean_email(v: Any) -> str:
            s = str(v).strip()
            # 去掉 mailto: 前缀
            s = re.sub(r"^mailto:", "", s, flags=re.IGNORECASE)
            # 去掉多余的 <>
            s = s.strip("<>")
            return s
        out["contact_email"] = out["contact_email"].apply(
            lambda x: clean_email(x) if pd.notna(x) and str(x).strip() else str(x)
        )

    # 手机号标准化
    if "contact_phone" in out.columns:
        def clean_phone(v: Any) -> str:
            s = str(v).strip()
            s = re.sub(r"\s+", "", s)  # 去空格
            s = re.sub(r"^tel:", "", s, flags=re.IGNORECASE)
            # 统一国际格式前导
            s = s.replace("+86-", "+86").replace("（", "(").replace("）", ")")
            return s
        out["contact_phone"] = out["contact_phone"].apply(
            lambda x: clean_phone(x) if pd.notna(x) and str(x).strip() else str(x)
        )

    return out


def _validate_field_values(df: pd.DataFrame) -> pd.DataFrame:
    """校验字段值合法性，清除明显无效的数据。"""
    out = df.copy()

    # contact_email 必须含 @
    if "contact_email" in out.columns:
        def valid_email(v: Any) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if not s or s in ("nan", "None", ""):
                return ""
            if "@" not in s or len(s) < 6:
                return ""
            if any(s.lower().endswith(ext) for ext in (".jpg", ".png", ".gif", ".css", ".js", ".pdf")):
                return ""
            return s
        out["contact_email"] = out["contact_email"].apply(valid_email)

    # company_name 不能是纯数字（线索编号等）
    if "company_name" in out.columns:
        def valid_name(v: Any) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if not s or s in ("nan", "None", ""):
                return ""
            # 纯数字 → 清空（后续 runner.py 会从 notes 中重新推断）
            if re.match(r"^\d{4,}$", s):
                return ""
            # URL → 清空（避免网址被当成公司名）
            if s.lower().startswith(("http://", "https://", "www.")):
                return ""
            return s
        out["company_name"] = out["company_name"].apply(valid_name)

    # contact_phone 至少 7 位数字
    if "contact_phone" in out.columns:
        def valid_phone(v: Any) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if not s or s in ("nan", "None", ""):
                return ""
            digits = re.sub(r"\D", "", s)
            if len(digits) < 7:
                return ""
            return s
        out["contact_phone"] = out["contact_phone"].apply(valid_phone)

    # contact_name 不能是纯数字
    if "contact_name" in out.columns:
        def valid_contact(v: Any) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if not s or s in ("nan", "None", ""):
                return ""
            if re.match(r"^\d+$", s):
                return ""
            return s
        out["contact_name"] = out["contact_name"].apply(valid_contact)

    return out


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """智能清洗 DataFrame：空行列 → 表头检测 → 合并单元格 → 格式标准化。"""
    logger.info("开始数据清洗: shape=%s", df.shape)

    # 1. 删除全空行列
    df = _strip_empty_rows_cols(df)
    logger.info("清理空行列后: shape=%s", df.shape)

    if df.empty:
        return df

    # 2. 处理合并单元格
    df = _handle_merged_cells(df)

    # 3. 表头检测
    header_row = _detect_header_row(df)
    if header_row > 0:
        # 前 header_row 行不是表头，设为列名并从下一行开始
        new_cols = [str(df.iloc[header_row, i]) for i in range(df.shape[1])]
        df = df.iloc[header_row + 1:].reset_index(drop=True)
        df.columns = new_cols
        logger.info("检测到表头在第 %d 行，重置列名", header_row + 1)
    elif header_row < 0:
        # 无法检测到表头——但如果当前列名像合法表头（pd.read_excel 已用第一行当列名），保留即可
        current_names = [str(c) for c in df.columns]
        if _is_header_row(current_names):
            logger.info("未检测到嵌入表头，但当前列名已是合法表头，保留")
        else:
            df.columns = [f"col_{i}" for i in range(df.shape[1])]
            logger.info("未检测到表头，使用默认列名 col_0..col_%d", df.shape[1] - 1)

    # 4. 去除全空行（可能来自表头分离后的残留）
    df = _strip_empty_rows_cols(df)

    # 5. 标准化格式
    df = _standardize_contact_formats(df)

    # 6. 校验字段值（清除纯数字公司名、无@邮箱等）
    df = _validate_field_values(df)

    logger.info("数据清洗完成: shape=%s, columns=%s", df.shape, list(df.columns))
    return df


# ============================================================
#  CSV 读取
# ============================================================

def _detect_csv_encoding(filepath: Path) -> str:
    """检测 CSV 文件编码，依次尝试常见编码。"""
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030", "latin-1", "iso-8859-1", "cp1252"]
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"  # 最后兜底


def _detect_csv_separator(filepath: Path, encoding: str) -> str:
    """检测 CSV 文件的分隔符。"""
    with open(filepath, "r", encoding=encoding) as f:
        first_lines = "".join(f.readline() for _ in range(5))
    candidates = {"\t": 0, ",": 0, ";": 0, "|": 0}
    for line in first_lines.strip().split("\n"):
        for sep in candidates:
            candidates[sep] += line.count(sep)
    best = max(candidates, key=candidates.get)
    if candidates[best] == 0:
        return ","
    return best


def read_input_csv(path: Path, *, meta: dict[str, Any] | None = None) -> tuple[pd.DataFrame, list[str]]:
    """读取 CSV 文件，自动检测编码和分隔符，清洗后走与 Excel 相同的后续流程。"""
    meta = meta or load_excel_io()
    encoding = _detect_csv_encoding(path)
    sep = _detect_csv_separator(path, encoding)
    logger.info("CSV 检测: encoding=%s, sep=%r", encoding, sep)

    # 先用 pandas 尝试读取
    try:
        df = pd.read_csv(path, encoding=encoding, sep=sep, dtype=str)
    except Exception:
        # 回退：让 pandas 自行嗅探
        df = pd.read_csv(path, encoding=encoding, sep=None, engine="python", dtype=str)

    logger.info("CSV 原始: shape=%s", df.shape)

    # 清洗
    df = _clean_dataframe(df)

    # 走与 Excel 相同的列映射流程
    canon = canonical_input_names(meta)
    cmap = normalize_column_map([str(c) for c in df.columns], canon)
    df = df.rename(columns=cmap)
    df = _dedup_columns(df)
    for c in meta["input_columns"]:
        if c["name"] in df.columns:
            continue
        if not c.get("required"):
            df[c["name"]] = ""
    df = _maybe_fill_notes_from_contacts(df)
    df = merge_extra_columns_into_notes(df, meta)
    missing = [c["name"] for c in meta["input_columns"] if c.get("required") and c["name"] not in df.columns]
    return df, missing


def load_excel_io(path: Path | None = None) -> dict[str, Any]:
    p = path or SCHEMA_EXCEL_IO
    return json.loads(p.read_text(encoding="utf-8"))


def canonical_input_names(meta: dict[str, Any]) -> list[str]:
    return [c["name"] for c in meta["input_columns"]]


def _maybe_fill_notes_from_contacts(df: pd.DataFrame) -> pd.DataFrame:
    """notes 为空时，不自动生成备注，改为保持空值。
    联系人信息已存放在独立列中，无需拼入 notes 造成杂乱。
    """
    return df  # 不再自动生成备注


def _fuzzy_match_column(col: str, canon: list[str], threshold: float = 0.55) -> str | None:
    """模糊匹配列名：先用关键词匹配，再用编辑距离兜底。"""
    s = col.strip().lower()

    # 1. 关键词匹配（中文子串命中优先）
    for keyword, target in _COLUMN_KEYWORDS.items():
        if keyword in s:
            if target in canon:
                return target

    # 2. 精确别名匹配（已在 normalize_column_map 中处理过了，此处为额外兜底）
    if col in COLUMN_ALIASES_TO_CANON:
        return COLUMN_ALIASES_TO_CANON[col]

    # 3. 编辑距离兜底（只对英文字符串）
    if not re.search(r"[一-鿿]", s):
        candidates = [(c, difflib.SequenceMatcher(None, s, c.lower()).ratio()) for c in canon]
        best = max(candidates, key=lambda x: x[1], default=("", 0))
        if best[1] >= threshold:
            return best[0]

    # 4. 中文列名用编辑操作比例
    best_canon = None
    best_score = 0.0
    for c in canon:
        score = difflib.SequenceMatcher(None, col, c).ratio()
        if score > best_score:
            best_score = score
            best_canon = c
    if best_canon and best_score >= threshold + 0.15:  # 更严格阈值
        return best_canon

    return None


def normalize_column_map(columns: list[str], canonical: list[str]) -> dict[str, str]:
    """原始列名 -> 规范列名。精确匹配优先于模糊匹配，避免数据列被误映射。"""
    lower_to_canon = {c.lower(): c for c in canonical}
    m: dict[str, str] = {}
    used_targets: set[str] = set()

    # 第一遍：精确匹配（别名或大小写不敏感）
    for col in columns:
        key = col.strip()
        target = None
        if key in COLUMN_ALIASES_TO_CANON:
            target = COLUMN_ALIASES_TO_CANON[key]
        elif key.lower() in lower_to_canon:
            target = lower_to_canon[key.lower()]
        if target and target not in used_targets:
            m[col] = target
            used_targets.add(target)
        elif target:
            logger.warning("列 '%s' 映射到 '%s'，但该目标已被占用，跳过", col, target)

    # 第二遍：未匹配的列才尝试模糊匹配
    for col in columns:
        if col in m:
            continue
        key = col.strip()
        target = _fuzzy_match_column(key, canonical)
        if target and target not in used_targets:
            m[col] = target
            used_targets.add(target)
        elif target:
            logger.warning("列 '%s' 模糊匹配到 '%s'，但该目标已被占用，跳过", col, target)
    return m


def read_input_xlsx(path: Path, *, meta: dict[str, Any] | None = None) -> tuple[pd.DataFrame, list[str]]:
    meta = meta or load_excel_io()
    canon = canonical_input_names(meta)
    df = pd.read_excel(path, engine="openpyxl")
    # 智能清洗
    df = _clean_dataframe(df)
    cmap = normalize_column_map([str(c) for c in df.columns], canon)
    df = df.rename(columns=cmap)
    df = _dedup_columns(df)
    for c in meta["input_columns"]:
        if c["name"] in df.columns:
            continue
        if not c.get("required"):
            df[c["name"]] = ""
    df = _maybe_fill_notes_from_contacts(df)
    df = merge_extra_columns_into_notes(df, meta)
    missing = [c["name"] for c in meta["input_columns"] if c.get("required") and c["name"] not in df.columns]
    return df, missing


def _try_extract_website_from_text(text: str) -> str | None:
    """从一段文本中提取最可能的公司网址。"""
    if not text:
        return None
    # 匹配 URL 模式
    urls = re.findall(r'(?:https?://)?(?:www\.)?[a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(?:/[^\s，。,；;|\n]*)?', text)
    if not urls:
        return None
    # 过滤常见无关域名
    skip_domains = {"example.com", "domain.com", "alibaba.com", "linkedin.com", "facebook.com",
                    "twitter.com", "youtube.com", "instagram.com", "google.com", "gmail.com",
                    "yahoo.com", "hotmail.com", "outlook.com", "163.com", "qq.com", "126.com"}
    for u in urls:
        u = u.strip().rstrip("/.")
        if not u.startswith("http"):
            u = "https://" + u
        # 提取域名
        domain = re.sub(r'^https?://(?:www\.)?', '', u).split('/')[0].lower()
        if domain not in skip_domains and not domain.endswith(('.jpg', '.png', '.pdf', '.css', '.js')):
            return u
    return None


def merge_extra_columns_into_notes(df: pd.DataFrame, meta: dict[str, Any]) -> pd.DataFrame:
    """未映射列优先匹配到规范字段（website/email/phone 等），其余才并入 notes。"""
    canon = set(canonical_input_names(meta))
    extras = [str(c) for c in df.columns if str(c) not in canon]
    if not extras:
        return df

    # 对每个额外列，尝试映射到有用的规范字段
    extra_to_target: dict[str, str] = {}
    for c in extras:
        cl = c.lower()
        # 尝试匹配 website
        if any(kw in cl for kw in ("网址", "网站", "url", "website", "domain", "官网", "网页")):
            extra_to_target[c] = "website"
        elif any(kw in cl for kw in ("邮箱", "email", "e-mail", "邮件")):
            extra_to_target[c] = "contact_email"
        elif any(kw in cl for kw in ("电话", "phone", "tel", "手机", "固话")):
            extra_to_target[c] = "contact_phone"
        elif any(kw in cl for kw in ("地址", "address")):
            extra_to_target[c] = "contact_address"
        elif any(kw in cl for kw in ("联系人", "姓名", "contact", "name")):
            extra_to_target[c] = "contact_name"
        elif any(kw in cl for kw in ("公司名", "公司", "企业", "company", "客户名")):
            extra_to_target[c] = "company_name"
        elif any(kw in cl for kw in ("国家", "country", "地区")):
            extra_to_target[c] = "country_region"

    out = df.copy()

    # 回填：将匹配到的额外列值写入目标列（当目标列为空时）
    for extra_col, target_col in extra_to_target.items():
        if target_col not in out.columns:
            out[target_col] = ""
        # 只在目标列为空时才覆盖
        for idx in out.index:
            target_val = out.at[idx, target_col]
            if target_val is None or (isinstance(target_val, float) and pd.isna(target_val)) or str(target_val).strip() == "":
                v = out.at[idx, extra_col]
                if v is not None and not (isinstance(v, float) and pd.isna(v)):
                    out.at[idx, target_col] = str(v).strip()

    # 剩余未映射的额外列 → 拼入 notes
    unmapped = [c for c in extras if c not in extra_to_target]

    def augment(row: pd.Series) -> str:
        parts: list[str] = []
        cur = row.get("notes")
        if cur is not None and not (isinstance(cur, float) and pd.isna(cur)):
            s0 = str(cur).strip()
            if s0:
                parts.append(s0)
        for c in unmapped:
            v = row.get(c)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            s = str(v).strip()
            if s:
                parts.append(f"{c}: {s}")
        return " | ".join(parts)

    if unmapped:
        out["notes"] = out.apply(augment, axis=1)

    # 从 notes 中尝试提取 website（当 website 列仍为空时）
    if "website" in out.columns:
        for idx in out.index:
            w = out.at[idx, "website"]
            if w is None or (isinstance(w, float) and pd.isna(w)) or str(w).strip() == "":
                notes_val = out.at[idx, "notes"]
                extracted = _try_extract_website_from_text(str(notes_val) if notes_val else "")
                if extracted:
                    out.at[idx, "website"] = extracted

    return out


def output_column_names(meta: dict[str, Any]) -> list[str]:
    return [c["name"] for c in meta["output_columns_append"]]


def ensure_output_columns(df: pd.DataFrame, meta: dict[str, Any]) -> pd.DataFrame:
    """按 excel_io 中的 type 初始化列，避免整列填 \"\" 导致 pandas 3 字符串 dtype 无法接受整数评分。"""
    idx = df.index
    for col in meta.get("output_columns_append") or []:
        name = col["name"]
        if name in df.columns:
            continue
        typ = col.get("type", "string")
        if typ == "int":
            df[name] = pd.Series(pd.NA, index=idx, dtype="Int64")
        elif typ == "float":
            df[name] = pd.Series(np.nan, index=idx, dtype="float64")
        else:
            df[name] = ""
    return df


def deal_recommendation_display_zh(raw: str) -> str:
    s = str(raw or "").strip().lower()
    if s == "high_intent":
        return "高意向跟进"
    if s == "watch":
        return "观察"
    if s == "no":
        return "不建议深入"
    return raw or ""


def buyer_seller_role_display_zh(raw: str) -> str:
    """LLM 枚举 buyer/seller/both/unclear → Summary 中文列。"""
    s = str(raw or "").strip().lower()
    if s == "buyer":
        return "买方"
    if s == "seller":
        return "卖方"
    if s == "both":
        return "兼营/难区分"
    if s == "unclear":
        return "不明"
    return raw or ""


def social_profiles_display_zh(raw: str) -> str:
    """将 JSON 数组转为可读的社交媒体账号摘要。"""
    if not raw:
        return ""
    try:
        profiles = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return raw
    if not isinstance(profiles, list):
        return raw
    parts = []
    for p in profiles:
        platform = p.get("platform", "")
        handle = p.get("handle", "")
        if platform and handle:
            parts.append(f"{platform}: {handle}")
    return "; ".join(parts)


def summary_export_spec(meta: dict[str, Any]) -> list[tuple[str, str]]:
    """(内部字段名, Excel 列标题)。"""
    spec = meta.get("summary_export_columns")
    if not isinstance(spec, list) or not spec:
        return [
            ("_source_row_1", "数据行号"),
            ("company_name", "客户名称"),
            ("website", "网站"),
            ("country_region", "国家"),
            ("contact_address", "联系人地址"),
            ("contact_phone", "固定电话"),
            ("contact_email", "联系人邮箱"),
            ("contact_name", "联系人姓名"),
            ("buyer_seller_role_display", "买方/卖方"),
            ("buyer_seller_reason", "角色判断依据"),
            ("product_fit_reasons", "产品匹配说明"),
            ("product_fit_score", "产品匹配分"),
            ("capability_score", "能力评分"),
            ("capability_signals", "能力依据"),
            ("reputation_facts", "资信要点"),
            ("reputation_safety_score", "资信安全分"),
            ("deal_recommendation_display", "合作建议"),
            ("manual_review_flag", "需复核"),
            ("overall_score_computed", "综合分"),
            ("social_profiles_display", "社交媒体"),
        ]
    out: list[tuple[str, str]] = []
    for item in spec:
        if isinstance(item, dict) and item.get("field") and item.get("header"):
            out.append((str(item["field"]), str(item["header"])))
    return out


def _cell_export(v: Any) -> Any:
    if v is None:
        return ""
    # Guard against pandas Series (duplicate column names)
    if hasattr(v, "iloc") and hasattr(v, "shape"):
        try:
            if len(v) > 0:
                v = v.iloc[0]
            else:
                return ""
        except Exception:
            pass
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return v


def build_summary_export_df(
    df: pd.DataFrame,
    meta: dict[str, Any],
    *,
    source_row_1based: list[int] | None = None,
) -> pd.DataFrame:
    """从内部宽表生成对外 Summary 数据框（列由 summary_export_columns 定义）。

    ``source_row_1based`` 与 ``df`` 行一一对应；用于分批导出时标注原始输入表行号（1-based）。
    未传时 ``_source_row_1`` 列默认为当前导出块内序号 1..n。
    """
    spec = summary_export_spec(meta)
    int_fields = {"product_fit_score", "capability_score", "reputation_safety_score"}
    rows: list[dict[str, Any]] = []
    for j, (_, row) in enumerate(df.iterrows()):
        out_row: dict[str, Any] = {}
        for field, header in spec:
            if field == "deal_recommendation_display":
                raw = row.get("deal_recommendation", "")
                val = deal_recommendation_display_zh(str(raw))
            elif field == "buyer_seller_role_display":
                raw = row.get("buyer_seller_role", "")
                val = buyer_seller_role_display_zh(str(raw))
            elif field == "social_profiles_display":
                raw = row.get("social_profiles", "")
                val = social_profiles_display_zh(str(raw))
            elif field == "_source_row_1":
                if source_row_1based is not None and j < len(source_row_1based):
                    val = int(source_row_1based[j])
                else:
                    val = j + 1
            else:
                v = _cell_export(row.get(field, ""))
                if field == "overall_score_computed":
                    if v == "":
                        val = ""
                    else:
                        try:
                            val = round(float(v), 3)
                        except (TypeError, ValueError):
                            val = v
                elif field in int_fields and v != "":
                    try:
                        val = int(float(v))
                    except (TypeError, ValueError):
                        val = v
                else:
                    val = v
            out_row[header] = val
        rows.append(out_row)
    return pd.DataFrame(rows, columns=[h for _, h in spec])


MAIN_SHEET_NAME = "Summary"


def write_result_xlsx(
    df: pd.DataFrame,
    path: Path,
    *,
    detail_df: pd.DataFrame | None = None,
    highlight_manual_review: bool = True,
    main_sheet_name: str = MAIN_SHEET_NAME,
    highlight_summary_row_indices: set[int] | None = None,
) -> None:
    """写出 Summary；可选 Detail；并对需复核行做整行浅红填充。

    highlight_summary_row_indices: 0-based 行号（与 df 行对齐），用于 Summary 无 manual_review 列时仍可按内部规则着色。
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # 清理 Excel 非法字符
        df = df.map(_sanitize_excel_text)
        df.to_excel(writer, sheet_name=main_sheet_name, index=False)
        if detail_df is not None and not detail_df.empty:
            detail_df = detail_df.map(_sanitize_excel_text)
            detail_df.to_excel(writer, sheet_name="Detail", index=False)

    if not highlight_manual_review:
        return

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(path)
    if main_sheet_name not in wb.sheetnames:
        return
    ws = wb[main_sheet_name]
    fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    headers = [c.value for c in ws[1]]
    flag_header = None
    for name in ("需复核", "manual_review_flag"):
        if name in headers:
            flag_header = name
            break

    if flag_header is not None:
        col_idx = headers.index(flag_header) + 1
        for row in range(2, ws.max_row + 1):
            flag_cell = ws.cell(row=row, column=col_idx)
            val = str(flag_cell.value or "").strip().upper()
            if val == "YES":
                for c in ws[row]:
                    c.fill = fill
        wb.save(path)
        return

    if highlight_summary_row_indices is not None:
        for i in highlight_summary_row_indices:
            excel_row = i + 2
            if 2 <= excel_row <= ws.max_row:
                for c in ws[excel_row]:
                    c.fill = fill

    wb.save(path)
